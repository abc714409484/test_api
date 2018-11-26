# -*- conding:utr-8 -*-
#@Time  :2018/11/17 14:05
#@Author:GYP测试
#@File  :do_excel.py
from openpyxl import load_workbook
from tools.do_config import ReadConfig
from tools.project_path import *

class Do_Excle:
    def __init__(self):
        self.file_name = test_data_path
        self.sheet_names = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
    def Read_Excle(self):
        wb = load_workbook(self.file_name)
        tel=15096090552
        test_data = []
        for sheet_name in self.sheet_names:
            sheet = wb[sheet_name]
            if self.sheet_names[sheet_name]=='all':
                for i in range(2,sheet.max_row+1):
                    sub_data={}
                    sub_data['case_id']=sheet.cell(i,1).value
                    sub_data['module']=sheet.cell(i,2).value
                    sub_data['description']=sheet.cell(i,3).value
                    sub_data['url']=sheet.cell(i,4).value
                    # sub_data['data']=sheet.cell(i,5).value
                    if sheet.cell(i,5).value.find('${tel_1}')!=-1:
                        sub_data['data']=sheet.cell(i,5).value.replace('${tel_1}',str(tel))
                    else:
                        sub_data['data'] = sheet.cell(i, 5).value

                    sub_data['method']=sheet.cell(i,6).value
                    sub_data['ExpectedResult']=sheet.cell(i,7).value
                    sub_data['ActaulResult']=sheet.cell(i,8).value
                    test_data.append(sub_data)
            else:
                for case_id in self.sheet_names[sheet_name]:
                    sub_data={}
                    sub_data['case_id'] = sheet.cell(case_id+1, 1).value
                    sub_data['module'] = sheet.cell(case_id+1, 2).value
                    sub_data['description'] = sheet.cell(case_id+1, 3).value
                    sub_data['url'] = sheet.cell(case_id+1, 4).value
                    # sub_data['data'] = sheet.cell(case_id+1, 5).value
                    if sheet.cell(i,5).value.find('${tel_1}')!=-1:
                        sub_data['data']=sheet.cell(i,5).value.replace('${tel_1}',tel)
                    else:
                        sub_data['data'] = sheet.cell(i, 5).value
                    sub_data['method'] = sheet.cell(case_id+1, 6).value
                    sub_data['ExpectedResult'] = sheet.cell(case_id+1, 7).value
                    sub_data['ActaulResult'] = sheet.cell(case_id+1, 8).value
                    test_data.append(sub_data)
        return test_data
    @staticmethod
    def write_excel(fiel_name,sheet_name,i,ActaulResult):
        wb=load_workbook(fiel_name)
        sheet=wb[sheet_name]
        sheet.cell(i,8).value=ActaulResult
        wb.save(fiel_name)
if __name__ == '__main__':
    res=Do_Excle().Read_Excle()
    print(res)
    print(len(res))

